# Mode: After MSFragger
import os,re
import sys
import pandas as pd
import openpyxl
from pathlib import Path
from AnalysisSignalModule import AnalysisSignal
import logging
from math import log,log2
from collections import defaultdict
# Define a function to find tsv files
analysis_signal = AnalysisSignal()
class Analysis:
    @staticmethod
    def find_index_with_regex(string_list, pattern):
        """
        在字符串列表中使用正则表达式查找匹配的元素，并返回其索引。

        Args:
        - string_list (list): 包含字符串的列表。
        - pattern (str): 要匹配的正则表达式模式。

        Returns:
        - int or None: 匹配的元素的索引，如果未找到匹配项则返回None。
        """
        for index, string in enumerate(string_list):
            if re.search(pattern, string):
                return index
        return None  # 如果没有找到匹配项，则返回None。
    @staticmethod
    # Define a function to find repeat elements in a single list
    def find_duplicates(lst):
        element_dict = {}  # 用于存储元素及其索引的字典
        duplicates = {}  # 存储重复元素及其所有索引的字典
        unique_elements_dict = {}  # 存储不重复元素及其索引的字典
        for i, item in enumerate(lst):
            # 如果元素已经在字典中，说明是重复元素
            if item in element_dict:
                if item not in duplicates:
                    duplicates[item] = [element_dict[item]]
                # 检查当前索引是否已经在列表中
                if i not in duplicates[item]:
                    duplicates[item].append(i)
                    # 从不重复字典中删除已经出现在重复字典中的元素
                    if item in unique_elements_dict:
                        del unique_elements_dict[item]
            else:
                # 将元素及其索引添加到字典中
                element_dict[item] = i
                # 如果元素已经在不重复字典中，不再更新其索引
                if item not in unique_elements_dict:
                    unique_elements_dict[item] = i
        return duplicates, unique_elements_dict
    def __init__(self,file_psm1,file_psm2,Save_path,fasta_input,Multiple_selection,Mode):
        d1 = file_psm1
        d2 = file_psm2
        print(d1)
        print(d2)
        self.Save = Path(Save_path)
        self.d = {1:d1,2:d2}
        self.Both_Peptide = {}
        self.folder_save_after = {}
        self.Both_ID_list = []
        self.Ratio3 = {}
        self.fo = []
        self.Intensity_final_use = {}
        self.both_intensity_all = {}
        self.Mode = Mode
        self.fasta_input = fasta_input
        self.Multiple_selection = Multiple_selection
        # 将输出重定向到AnalysisSignal
        analysis_signal.write(text=f"Starting analysis with \nPSM Group1: {file_psm1}, \nPSM Group2: {file_psm2}, \nSave path: {Save_path}")
        logging.info(f"Starting analysis with \nPSM1: {file_psm1}, \nPSM2: {file_psm2}, \nSave path: {Save_path}")

    def run(self):
        # the main CODE in the whole program
        for i,value_in in self.d.items():
            self.Both_Peptide[i] = {}
            self.both_intensity_all[i] = {}
            self.folder_save_after[i] = {}
            for ifl,f in enumerate(value_in):
                # print(i)
                file_path = f
                # print(file_path)
                form = pd.read_csv(file_path, sep='\t', header=0)

                workbook = openpyxl.Workbook(write_only=False)
                # Get the parent directories of each file
                parent_directories = (os.path.dirname(file_path))
                # Remove characters before and including the last part of each directory
                modified_directories = parent_directories.split("/")[-1]
                invalid_chars = r'[\\/*?:\[\]]'
                modified_dir = modified_directories
                modified_dir = re.sub(invalid_chars, '', modified_dir)
                counter = 1
                while True:
                    folder = os.path.join(self.Save, f"{modified_dir}_MSFragger_{counter}").replace("/", "\\")
                    self.folder_save_after[i][ifl] = folder
                    if not os.path.exists(folder):
                        os.makedirs(folder)
                        excel_file_path = f'{folder}\\output_Group{i}_{ifl+1}_{modified_dir}.xlsx'
                        self.fo.append(folder)
                        sheet_name = modified_dir
                    # 一次性进行数据筛选
                        if self.Mode == "N-Glycopeptide":
                            filtered_data = form[(form['Intensity'] != 0)]
                        elif self.Mode == "O-Glycopeptide":
                            filtered_data = form[(form['Intensity'] != 0)]

                    # 在工作簿中创建一个新的工作表

                        sheet = workbook.create_sheet(sheet_name)
                        # 写入表头
                        if self.Mode == "N-Glycopeptide":
                            sheet.append(['Protein_ID', 'Peptide', 'Intensity', 'MSFragger Localization'])
                        elif self.Mode == "O-Glycopeptide":
                            sheet.append(['Protein_ID', 'Peptide', 'Intensity'])

                        Peptide = filtered_data['Peptide'].reset_index(drop=True)
                        # Log-transform the intensity values
                        filtered_data['Intensity'] = filtered_data['Intensity']
                        Intensity = filtered_data['Intensity'].reset_index(drop=True)

                        self.Intensity_final_use[i] = Intensity
                        Protein_ID = filtered_data['Protein ID'].reset_index(drop=True)
                        # 逐行写入数据
                        if self.Mode == "N-Glycopeptide":
                            for row_data in zip(filtered_data['Protein ID'], filtered_data['Peptide'], filtered_data['Intensity'], filtered_data['MSFragger Localization']):
                                sheet.append(row_data)
                        elif self.Mode == "O-Glycopeptide":
                            for row_data in zip(filtered_data['Protein ID'], filtered_data['Peptide'], filtered_data['Intensity']):
                                sheet.append(row_data)
                        # print(Peptide)
                        # print(type(Peptide))
                        # 格式调整
                        Peptide_Analysis = workbook.create_sheet('Peptide_Analysis')
                        peptide_dealing_R,peptide_dealing_NR = self.find_duplicates(Peptide.tolist())
                        # for item in peptide_dealing_R:
                        # print(peptide_dealing_R)
                        # print(peptide_dealing_NR)
                        # 删除默认的Sheet
                        default_sheet = workbook['Sheet']
                        workbook.remove(default_sheet)
                        Total_peptide = []

                        Rpart = {}
                        # Calculate the sum of all intensities for both repeating and non-repeating peptides
                        # 确保 Intensity 列是数值型
                        filtered_data['Intensity'] = pd.to_numeric(filtered_data['Intensity'], errors='coerce')
                        # 计算总和
                        total_intensity_all = filtered_data['Intensity'].sum()
                        self.both_intensity_all[i][ifl] = total_intensity_all
                        # Loop through peptide_dealing_R
                        for key, values in peptide_dealing_R.items():
                            for value in values:
                                ID = str(Protein_ID.loc[value])
                                intensity = float(Intensity.loc[value])
                                # Use a tuple (ID, key) as a key for the dictionary
                                key_tuple = (ID, key)
                                # If the key is not in the dictionary, add it with the current intensity
                                if key_tuple not in Rpart:
                                    Rpart[key_tuple] = intensity
                                else:
                                    # If the key is already in the dictionary, update the intensity
                                    Rpart[key_tuple] += intensity
                        # Convert the dictionary to a list of lists
                        part_list1 = [[ID, key, intensity] for (ID, key), intensity in Rpart.items()]
                        Total_peptide.append(part_list1)
                        # print(part_list1)
                        # Write data to 'Peptide_Analysis' sheet for repeating peptides
                        Peptide_Analysis.append(['Protein ID', 'Peptide', 'Total Intensity', 'Ratio'])
                        for row_data in part_list1:
                            # Calculate the ratio for each row
                            ratio = row_data[2] / total_intensity_all
                            Peptide_Analysis.append(row_data + [ratio])
                        # Loop through peptide_dealing_NR
                        part_list2 = []
                        for key, value in peptide_dealing_NR.items():
                            ID = str(Protein_ID.loc[value])
                            intensity = float(Intensity.loc[value])
                            # Calculate the ratio for each non-repeating peptide
                            ratio = intensity / total_intensity_all
                            # Append data to 'part_list2' list for non-repeating peptides
                            part_list2.append([ID,key,intensity,ratio])
                        # Write data to 'Peptide_Analysis' sheet for non-repeating peptides
                        for row_data in part_list2:
                            Peptide_Analysis.append(row_data)
                        # print(part_list2)
                        Total_peptide.append(part_list2)
                        self.Both_Peptide[i][ifl] = Total_peptide
                    # print(Both_Peptide)

                    # 保存Excel文件
                    # 设置自适应列宽(Sheet)
                        for column in sheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                        pass
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
                        # 设置自适应列宽(Peptide_Analysis)
                        for column in Peptide_Analysis.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            Peptide_Analysis.column_dimensions[column[0].column_letter].width = adjusted_width
                        # print(folder)

                        workbook.save(excel_file_path)
                        workbook.close()
                        print(f"Excel File has saved at: {excel_file_path}")
                        analysis_signal.write(text=f"Excel File has saved at: {excel_file_path}")
                        logging.info(f"Excel File has saved at: {excel_file_path}")
                        break

                    counter += 1

                # print(Total_peptide)
                ID_list = []
                Peptide_list = []
                Intensity_list = []
                for item in Total_peptide:  # Both_Peptide[i] = {ifl:Total_peptide} # 调整为Both_peptide输出
                    for tem_list in item:
                        tem_ID = tem_list[0]
                        ID_list.append(tem_ID)
                        tem_Peptide = tem_list[1]
                        Peptide_list.append(tem_Peptide)
                        tem_Intensity = tem_list[2]
                        Intensity_list.append(tem_Intensity)
                # Both_ID_list.append(ID_list)
                ID_list = tuple(ID_list)
                Peptide_list = tuple(Peptide_list)
                Intensity_list = tuple(Intensity_list)
                del Total_peptide
                # 输出重复元素与索引的函数部分，此处对元组进行操作
                indices = {}
                for index, item in enumerate(ID_list):
                    if item in indices:
                        indices[item].append(index)
                    else:
                        indices[item] = [index]
                # print(ID_list)
                # print(indices)
                import matplotlib.pyplot as plt
                import matplotlib as mpl
                from matplotlib import ticker
                mpl.rcParams['font.family'] = 'Times New Roman'
                x1_ID = []
                y1_Intensity_Ratio = []
                x2_Peptide = []
                for key in indices:
                    Intensity_Sum = 0
                    every_values = indices[key]
                    for item in every_values:
                        Intensity_Sum += Intensity_list[item]
                        x2_Peptide.append(Peptide_list[item])
                    x1_ID.append(key)
                    del total_intensity_all
                    total_intensity_all = self.both_intensity_all[i][ifl]
                    output = Intensity_Sum / total_intensity_all
                    y1_Intensity_Ratio.append(output)
                plt.rcParams['font.size'] = 12.5
                fig, ax = plt.subplots(figsize=(16, 9))  # 调整图表尺寸


                bars = ax.bar(x1_ID, y1_Intensity_Ratio)
                # self.Ratio3[(i, ifl)] = (x1_ID, y1_Intensity_Ratio)

                ax.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1, decimals=1))
                # 根据标签数量和图表宽度调整显示的标签数量
                total_labels = len(x1_ID)
                step = max(1,total_labels // 15)  # For example, a label is displayed every 15 points, which can be adjusted as needed
                plt.xticks(range(len(x1_ID))[::step], x1_ID[::step])
                for tick in ax.get_xticklabels():
                    tick.set_rotation(45)

                # 使用循环遍历数据点
                threshold = 0.1
                for bar,value in zip(bars, y1_Intensity_Ratio):
                    # 根据条件判断是否添加注释
                    if value > threshold:
                        ax.annotate(f'{value*100:.2f}%',  # 格式化数值
                                    xy=(bar.get_x() + bar.get_width() / 2, value),  # 获取柱形中心位置
                                    xytext=(0, 3),  # 调整文本位置
                                    textcoords="offset points",
                                    ha='center', va='bottom')
                plt.title("Protein Ratio")
                plt.autoscale(enable=True, axis='both', tight=None)
                plt.subplots_adjust(left=0.07, right=0.93, top=0.92, bottom=0.15)
                # plt.show()
                plt.savefig(f"{folder}\\Protein_Ratio_Group{i}_{ifl+1}_{modified_dir}.png", dpi=1000)
                plt.close()
        # Next Step is to finished the mutiple Ratio
        # workbook_c = openpyxl.Workbook(write_only=False)
        # workbook_c.active
        # sheet_c = workbook_c.create_sheet("Combination")
        # # 删除默认的Sheet
        # default_sheet = workbook_c['Sheet']
        # workbook_c.remove(default_sheet)

        # print(Both_Peptide)
        b = {}
        for i,value in self.Both_Peptide.items(): # Group Level
            # Both_Peptide[i] = {ifl:Total_peptide}
            tem2 = {}
            for i1,value_1 in value.items():  # File Level
                tem1 = {}
                ID_list = []
                Peptide_list = []
                Intensity_list = []
                for item in value_1:  # Repeat Peptide and Non-Repeat Peptide
                    for tem_list in item:
                        tem_ID = tem_list[0]
                        ID_list.append(tem_ID)
                        tem_Peptide = tem_list[1]
                        Peptide_list.append(tem_Peptide)
                        tem_Intensity = tem_list[2]
                        Intensity_list.append(tem_Intensity)
                ID_list = tuple(ID_list)  #  A File
                # R_ID_list,NR_ID_list = self.find_duplicates(ID_list)
                Intensity_list = tuple(Intensity_list)   #  A File
                # del Both_Peptide
                # 输出重复元素与索引的函数部分，此处对元组进行操作
                indices = {}
                for index, item in enumerate(ID_list):
                    if item in indices:
                        indices[item].append(index)
                    else:
                        indices[item] = [index]
                    # print(indices)
                    for key in indices:
                        Intensity_Sum = 0
                        every_values_ID = indices[key]
                        # every_values_Peptide = indices[key][1]
                        for item1 in every_values_ID:
                            Intensity_Sum += Intensity_list[item1]
                            # print(Intensity_Sum)
                        total_intensity_all = self.both_intensity_all[i][i1]
                        output1 = Intensity_Sum / total_intensity_all
                        tem1[key] = output1
                tem2[i1] = tem1
            b[i] = tem2
        print(b)
        del tem1,tem2
        tem3 = {}
        for key1,value1 in b.items():
            tem1 = []
            tem2 = []
            for key2,value2 in value1.items():
                for key3,value3 in value2.items():
                    tem1.append(key3)
            for s in tem1:
                if tem1.count(s) > 1:
                    tem2.append(s)
            # 取值，绘图
            for s in tem2:
                x = []
                y = []
                for key2,value2 in value1.items():
                    x.append('File'+str(key2 + 1))
                    l = value2.get(s,None)
                    if l == None:
                        l = 0
                    else:
                        pass
                    y.append(l)
                    tem3[(key1, key2, s)] = l
                # 20240417
                # mpl.rcParams['font.family'] = 'Times New Roman'
                # plt.rcParams['font.size'] = 12.5
                # fig, ax = plt.subplots(figsize=(18, 11))
                # plt.subplots_adjust(left=0.1, right=0.9, bottom=0.1, top=0.9)
                # bars = ax.bar(x, y)
                # ax.yaxis.set_major_formatter(ticker.PercentFormatter(xmax=1, decimals=1))
                # for bar, value in zip(bars, y):
                #     ax.annotate(f'{value * 100:.2f}%',  # 格式化数值
                #                 xy=(bar.get_x() + bar.get_width() / 2, value),  # 获取柱形中心位置
                #                 xytext=(0, 3),  # 调整文本位置
                #                 textcoords="offset points",
                #                 ha='center', va='bottom')
                # plt.title(f"Protein Ratio in Different Files-{s}")
                # ax.autoscale()
                # # plt.show()
                # Save_Protein_Group = os.path.join(Save, f"Image_Protein_Group{key1}").replace("/", "\\")
                # if not os.path.exists(Save_Protein_Group):
                #     os.makedirs(Save_Protein_Group)
                # plt.savefig(f"{Save_Protein_Group}\\Protein_Ratio in Different Files_Group{key1}_{s}.png", dpi=1000)
                # plt.close()
                # analysis_signal.write(f'Protein Different Value show at: {Save_Protein_Group}\\Protein_Ratio in Different Files_Group{key1}_{s}.png')
                # print(f'Protein Different Value show at: {Save_Protein_Group}\\Protein_Ratio in Different Files_Group{key1}_{s}.png')
        Ratio2 = []
        tem1_3 = None
        tem2_3 = None
        for key0, value0 in self.Both_Peptide.items():  # 循环两组
            tem1_1 = {}  # 文件内的ID和
            tem2_1 = {}  # 文件内的ID和
            tem1_2 = {}  # 文件内的ID比值
            tem2_2 = {}  # 文件内的ID比值
            for key1, value1 in value0.items():  # 循环组内各个文件
                ID_list = []
                Peptide_list = []
                Intensity_list = []
                for value2 in value1:  # 循环文件内重复和非重复的肽段数据
                    for tem_list in value2:
                        tem_ID = tem_list[0]
                        ID_list.append(tem_ID)
                        tem_Peptide = tem_list[1]
                        Peptide_list.append(tem_Peptide)
                        tem_Intensity = tem_list[2]
                        Intensity_list.append(tem_Intensity)
                R_ID, NR_ID = self.find_duplicates(ID_list)
                for ID, item in R_ID.items():
                    ID_sum = 0
                    for number in item:
                        ID_sum += float(Intensity_list[number])
                    if key0 == 1:
                        tem1_1[(key1, ID)] = (ID, ID_sum)
                    else:
                        tem2_1[(key1, ID)] = (ID, ID_sum)
                for ID, item in NR_ID.items():
                    ID_sum = 0
                    ID_sum += float(Intensity_list[item])
                    if key0 == 1:
                        tem1_1[(key1, ID)] = (ID, ID_sum)  # 问题在此
                    else:
                        tem2_1[(key1, ID)] = (ID, ID_sum)
                for key123, value123 in self.both_intensity_all.items():  # 两组总峰强
                    if key0 == 1:
                        # ID = []
                        # ratio = []
                        for key23, value23 in value123.items():  # 单个文件总峰强
                            for key124, value124 in tem1_1.items():  # 单个文件单一ID峰强
                                if key124[0] == key23:  # key124为文件号，key23为单个文件总峰强字典文件号
                                    average = value124[1] / value23
                                    tem1_2[(key23, key124[1])] = (key124[1], average)
                                    # ID.append(key124[1])
                                    # ratio.append(average)
                            # self.Ratio3[(1, key23)] = (ID,ratio)
                    else:
                        # ID = []
                        # ratio = []
                        for key23, value23 in value123.items():  # 单个文件总峰强
                            for key124, value124 in tem2_1.items():  # 单个文件单一ID峰强
                                if key124[0] == key23:  # key124为文件号，key23为单个文件总峰强字典文件号
                                    average = value124[1] / value23
                                    tem2_2[(key23, key124[1])] = (key124[1], average)
                                    # ID.append(key124[1])
                                    # ratio.append(average)
                            # self.Ratio3[(2, key23)] = (ID,ratio)
            # 对3个文件的相同ID进行取平均，第一组
            # 用于存储相同第一个值的元组的第二个值的列表
            value_dict = defaultdict(list)
            value_dict1 = defaultdict(list)
            if key0 == 1:
                for inner_key, (first, second) in tem1_2.items():
                    # 将第二个值添加到相同第一个值的列表中
                    value_dict[first].append(second)
                for list_key, values in value_dict.items():
                    if len(values) >= 3:  # 求组内均值时，分母是否需要严格取组内文件数量？ value0是上面循环的文件字典，value统计糖蛋白在文件内是否含有的字典 25.1.9 0：30
                        value_dict1[list_key] = values
                # 计算平均值并保存到新字典中
                average_dict = {key: sum(values) / len(values) for key, values in value_dict1.items()}

                r = {}
                for v1, v2 in value_dict1.items():
                    for indexr, raioi in enumerate(v2):
                        r[(key0, indexr, v1)] = raioi
                for (v0, indexr, v1), raioi in r.items():
                    new_key = (v0, indexr)
                    if new_key not in self.Ratio3:
                        self.Ratio3[new_key] = [[], []]
                    self.Ratio3[new_key][0].append(v1)
                    self.Ratio3[new_key][1].append(raioi)


                tem1_3 = average_dict  # 单组内的ID平均值
                del tem1_1, tem1_2
            # 对3个文件的相同ID进行取平均，第二组
            else:
                for inner_key, (first, second) in tem2_2.items():
                    # 将第二个值添加到相同第一个值的列表中
                    value_dict[first].append(second)
                for list_key, values in value_dict.items():
                    if len(values) >= 3: # 求组内均值时，分母是否需要严格取组内文件数量？ value0是上面循环的文件字典，value统计糖蛋白在文件内是否含有的字典 25.1.9 0：30
                        value_dict1[list_key]=values
                # 计算平均值并保存到新字典中
                average_dict = {key: sum(values) / len(values) for key, values in value_dict1.items()}

                r = {}
                for v1, v2 in value_dict1.items():
                    for indexr, raioi in enumerate(v2):
                        r[(key0, indexr, v1)] = raioi
                for (v0, indexr, v1), raioi in r.items():
                    new_key = (v0, indexr)
                    if new_key not in self.Ratio3:
                        self.Ratio3[new_key] = [[], []]
                    self.Ratio3[new_key][0].append(v1)
                    self.Ratio3[new_key][1].append(raioi)

                tem2_3 = average_dict  # 单组内的ID平均值
                del tem2_1, tem2_2
        for key1, value1 in tem1_3.items():
            for key2, value2 in tem2_3.items():
                if value1 != 0 and value2 != 0 and key1 == key2:
                    output = value1 / value2
                    if log2(output) >= log2(self.Multiple_selection):  # 上调
                        Ratio2.append((key1, output))
                        self.Both_ID_list.append(key1)
                    if log2(output) <= -log2(self.Multiple_selection):  # 下调 -0.585
                        Ratio2.append((key1, output))
                        self.Both_ID_list.append(key1)
        print(Ratio2)
        analysis_signal.write(text=Ratio2)
        logging.info(f"Ratio2:{Ratio2}")
        x1_ID = []
        y1_Intensity_Ratio = []
        for value1 in Ratio2:
            x1_ID.append(value1[0])
            y1_Intensity_Ratio.append(value1[1])
        jug = len(Ratio2)
        # jugement to make a decision to deal with the data
        if jug != 0:
            plt.rcParams['font.size'] = 12.5
            fig, ax = plt.subplots(figsize=(18, 11))
            plt.subplots_adjust(left=0.1, right=0.9, bottom=0.1, top=0.9)
            bars = ax.bar(x1_ID, y1_Intensity_Ratio)
            df = pd.DataFrame(({'Protein': x1_ID, 'Intens.': y1_Intensity_Ratio}))
            # X轴角度45°
            for tick in ax.get_xticklabels():
                tick.set_rotation(45)

            # 使用循环遍历数据点
            threshold = 0.1

            if jug > 15:
                step = jug // 15

                # 设置简要输出的X轴标签
                ax.set_xticks([i for i in range(0, jug, step)])
                ax.set_xticklabels([x1_ID[i] for i in range(0, jug, step)])

            for i, (bar, value) in enumerate(zip(bars, y1_Intensity_Ratio)):
                # 根据条件判断是否添加注释
                if value > threshold and len(y1_Intensity_Ratio) < 15:
                    ax.annotate('%.2f' % value,  # 格式化数值
                                xy=(bar.get_x() + bar.get_width() / 2, value),  # 获取柱形中心位置
                                xytext=(0, 3),  # 调整文本位置
                                textcoords="offset points",
                                ha='center', va='bottom')
            plt.title("Protein Ratio Difference Multiple")
            ax.autoscale()
            plt.savefig(f"{self.Save}\\Protein_Ratio_Difference.png", dpi=1000)
            plt.close()
            df.to_excel(f"{self.Save}\\ProteinRatio_Difference.xlsx", index=False)
            print("Done！")
            analysis_signal.write(text="Finished Compared Part")
            logging.info("Finished Compared Part")
            # print("Please press Enter to exit~~")
            # input()
        elif jug == 0:
            print("No relationship both of them......")
            analysis_signal.write(text="Both sets of data are the same\nFinished Compared Part!")
            logging.info("Both sets of data are the same\nFinished Compared Part!")
            # print("Please press Enter to exit~~")
            # input()
        try:
            list1 = set(self.Both_ID_list)
            list2 = list(list1)
            output_path = self.Save
            Total2 = list2
            fasta = self.fasta_input
            self.fasta_output = []
            with open(file=os.path.join(str(output_path) + "\\New_MSFragger_ProteinDatabase.fasta"), mode="w", encoding="utf-8") as r:
                for item1 in Total2:
                    for item2 in fasta:
                        op = re.search(re.escape(item1), item2)
                        if op:
                            matched_text = item2
                            output = matched_text
                            self.fasta_output.append(matched_text)
                            print(output)
                            r.write(output)
            r.close()
        except Exception as e:
            logging.info(e)
        return self.Ratio3, self.fasta_output
    # 恢复标准输出
    sys.stdout = sys.__stdout__
    analysis_signal.write("Compared Part of Analysis Thread completed.")
    logging.info("Compared Part of Analysis Thread completed.")